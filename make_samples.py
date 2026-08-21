# -*- coding: utf-8 -*-
"""
샘플 데이터 생성 스크립트
- js/seed.js               : 포털 시드 데이터(localStorage 초기값)
- templates/딜러실적_양식.xlsx : 딜러 대시보드 엑셀 업로드 빈 양식
- templates/딜러실적_샘플.xlsx : 샘플 값이 채워진 업로드 예시 파일

실행: python3 make_samples.py
"""
import json
import os
import random
from datetime import date, datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

random.seed(20260821)

BASE = os.path.dirname(os.path.abspath(__file__))
TODAY = date.today()

# ---------------------------------------------------------------------------
# 기준 기간 계산 (최근 12개월 / 최근 8주)
# ---------------------------------------------------------------------------

def month_label(d):
    return f"{d.year:04d}-{d.month:02d}"


def shift_month(d, n):
    y = d.year + (d.month - 1 + n) // 12
    m = (d.month - 1 + n) % 12 + 1
    return date(y, m, 1)


MONTHS = [month_label(shift_month(TODAY, -i)) for i in range(11, -1, -1)]  # 과거 -> 현재


def iso_week_label(d):
    y, w, _ = d.isocalendar()
    return f"{y:04d}-W{w:02d}"


WEEK_MONDAYS = []
monday = TODAY - timedelta(days=TODAY.weekday())
for i in range(7, -1, -1):
    WEEK_MONDAYS.append(monday - timedelta(weeks=i))
WEEKS = [iso_week_label(d) for d in WEEK_MONDAYS]  # 8주치, 과거 -> 현재

# ---------------------------------------------------------------------------
# 팀원 (사용자)
# ---------------------------------------------------------------------------
USERS = [
    {"id": "u1", "name": "최율아", "email": "yula.choi@example.com", "role": "책임자"},
    {"id": "u2", "name": "김도현", "email": "dohyun.kim@example.com", "role": "팀원"},
    {"id": "u3", "name": "박서연", "email": "seoyeon.park@example.com", "role": "팀원"},
    {"id": "u4", "name": "이준호", "email": "junho.lee@example.com", "role": "팀원"},
    {"id": "u5", "name": "정하늘", "email": "haneul.jung@example.com", "role": "팀원"},
    {"id": "u6", "name": "오민재", "email": "minjae.oh@example.com", "role": "팀원"},
]

# ---------------------------------------------------------------------------
# 딜러 14개 (가상명)
# ---------------------------------------------------------------------------
DEALERS = [
    {"code": "DL01", "name": "노르트몰 모터스", "country": "독일",      "region": "유럽법인",  "currency": "EUR", "manager": "김도현"},
    {"code": "DL02", "name": "알프스 트레이딩", "country": "스위스",    "region": "유럽법인",  "currency": "EUR", "manager": "김도현"},
    {"code": "DL03", "name": "이베리아 커머스", "country": "스페인",    "region": "유럽법인",  "currency": "EUR", "manager": "박서연"},
    {"code": "DL04", "name": "그레이트레이크 서플라이", "country": "미국", "region": "미주법인", "currency": "USD", "manager": "박서연"},
    {"code": "DL05", "name": "론스타 인더스트리", "country": "미국",    "region": "미주법인",  "currency": "USD", "manager": "이준호"},
    {"code": "DL06", "name": "아즈텍 파트너스", "country": "멕시코",    "region": "미주법인",  "currency": "USD", "manager": "이준호"},
    {"code": "DL07", "name": "카리오카 임포트", "country": "브라질",    "region": "미주법인",  "currency": "USD", "manager": "이준호"},
    {"code": "DL08", "name": "갠지스 오토", "country": "인도",          "region": "아시아법인", "currency": "USD", "manager": "정하늘"},
    {"code": "DL09", "name": "메콩 머시너리", "country": "베트남",      "region": "아시아법인", "currency": "USD", "manager": "정하늘"},
    {"code": "DL10", "name": "시암 이큅먼트", "country": "태국",        "region": "아시아법인", "currency": "USD", "manager": "정하늘"},
    {"code": "DL11", "name": "보르네오 리소스", "country": "인도네시아", "region": "아시아법인", "currency": "USD", "manager": "오민재"},
    {"code": "DL12", "name": "걸프스타 제너럴", "country": "UAE",       "region": "직수출",    "currency": "USD", "manager": "오민재"},
    {"code": "DL13", "name": "사바나 홀딩스", "country": "남아공",      "region": "직수출",    "currency": "USD", "manager": "오민재"},
    {"code": "DL14", "name": "안데스 커머셜", "country": "칠레",        "region": "직수출",    "currency": "USD", "manager": "박서연"},
]

TERMS = [
    "L/C at sight, 인코텀즈 FOB",
    "T/T 30일, 연간 최소 주문 200대",
    "T/T 60일, 분기 리베이트 2%",
    "D/A 90일, 독점 계약(국가 단위)",
    "L/C usance 60일, CIF 조건",
    "T/T 선급 30% + 잔금 70일",
]

# ---------------------------------------------------------------------------
# 딜러 실적 12개월 (단위: USD 천불 기준 환산)
# ---------------------------------------------------------------------------
metrics = []
for idx, dl in enumerate(DEALERS):
    base_sales = random.randint(180, 950)          # 월 매출 기본값(천불)
    trend = random.uniform(-0.015, 0.03)           # 월별 성장 추세
    base_inv = random.randint(120, 600)
    term = TERMS[idx % len(TERMS)]
    for mi, m in enumerate(MONTHS):
        season = 1 + 0.12 * random.uniform(-1, 1)
        sales = round(base_sales * (1 + trend) ** mi * season)
        inventory = max(40, round(base_inv + random.randint(-80, 80)))
        receivable = round(sales * random.uniform(0.9, 1.8))
        overdue = round(receivable * random.uniform(0.02, 0.22))
        metrics.append({
            "dealerCode": dl["code"],
            "month": m,
            "sales": sales,
            "inventory": inventory,
            "receivable": receivable,
            "overdue": overdue,
            "terms": term,
        })

# ---------------------------------------------------------------------------
# 주간업무 8주치 (4개 지역/법인)
# ---------------------------------------------------------------------------
REGIONS = ["유럽법인", "미주법인", "아시아법인", "직수출"]
REGION_AUTHOR = {"유럽법인": "u2", "미주법인": "u3", "아시아법인": "u5", "직수출": "u6"}
TASK_POOL = {
    "유럽법인": ["독일 딜러 신모델 론칭 지원", "스페인 재고 리밸런싱 협의", "유럽 인증(CE) 서류 갱신", "스위스 딜러 분기 실적 리뷰"],
    "미주법인": ["미국 동부 딜러 프로모션 정산", "멕시코 관세 변경 대응", "브라질 선적 일정 조율", "북미 A/S 부품 수급 점검"],
    "아시아법인": ["인도 신규 모델 가격 협상", "베트남 딜러 교육 진행", "태국 전시회 부스 준비", "인도네시아 재고 실사"],
    "직수출": ["UAE 신규 오더 계약 검토", "남아공 선적 서류 처리", "칠레 딜러 여신 한도 재심사", "중동 시장 경쟁사 동향 조사"],
}
ISSUE_POOL = [
    "환율 변동으로 마진 축소 우려",
    "선적 지연 2주 발생, 대체 선사 검토",
    "일부 모델 재고 부족",
    "경과채권 회수 지연 딜러 1곳 발생",
    "",
]
weekly = []
rid = 1
for wi, (wl, wm) in enumerate(zip(WEEKS, WEEK_MONDAYS)):
    for region in REGIONS:
        author = REGION_AUTHOR[region]
        sales = random.randint(120, 520)
        weekly.append({
            "id": f"w{rid}",
            "week": wl,
            "authorId": author,
            "region": region,
            "tasks": " / ".join(random.sample(TASK_POOL[region], 2)),
            "sales": sales,
            "issues": random.choice(ISSUE_POOL),
            "nextPlan": random.choice(TASK_POOL[region]),
            "createdAt": (wm + timedelta(days=4)).isoformat(),
        })
        rid += 1

# ---------------------------------------------------------------------------
# 환율 12개월 (KRW 기준)
# ---------------------------------------------------------------------------
CURRENCIES = {"USD": 1385.0, "EUR": 1512.0, "JPY": 9.35, "CNY": 191.0}
rates = []
xid = 1
for mi, m in enumerate(MONTHS):
    for cur, base in CURRENCIES.items():
        drift = base * (1 + random.uniform(-0.035, 0.035))
        rates.append({
            "id": f"x{xid}",
            "month": m,
            "currency": cur,
            "rate": round(drift, 2),
            "inputBy": "최율아",
            "inputAt": f"{m}-02",
        })
        xid += 1

# ---------------------------------------------------------------------------
# 출장 6건 / 회의 4건
# ---------------------------------------------------------------------------
def d(offset):
    return (TODAY + timedelta(days=offset)).isoformat()

trips = [
    {"id": "t1", "memberId": "u2", "countryCity": "독일 프랑크푸르트", "startDate": d(-9),  "endDate": d(-5), "purpose": "딜러 분기 실적 리뷰", "dealers": "DL01, DL02"},
    {"id": "t2", "memberId": "u3", "countryCity": "미국 시카고",       "startDate": d(-2),  "endDate": d(3),  "purpose": "신모델 론칭 미팅",   "dealers": "DL04"},
    {"id": "t3", "memberId": "u5", "countryCity": "베트남 하노이",     "startDate": d(4),   "endDate": d(8),  "purpose": "딜러 교육 및 재고 점검", "dealers": "DL09"},
    {"id": "t4", "memberId": "u6", "countryCity": "UAE 두바이",        "startDate": d(11),  "endDate": d(16), "purpose": "신규 오더 계약 협상", "dealers": "DL12"},
    {"id": "t5", "memberId": "u4", "countryCity": "멕시코 몬테레이",   "startDate": d(18),  "endDate": d(23), "purpose": "관세 대응 및 채권 회수 협의", "dealers": "DL06"},
    {"id": "t6", "memberId": "u1", "countryCity": "스페인 마드리드",   "startDate": d(25),  "endDate": d(29), "purpose": "유럽 딜러 총괄 미팅", "dealers": "DL03"},
]

meetings = [
    {"id": "m1", "datetime": d(-7) + "T09:30", "place": "본사 3층 회의실 A", "agenda": "주간 실적 리뷰 및 선적 일정 점검", "attendees": "전원"},
    {"id": "m2", "datetime": d(1) + "T09:30",  "place": "본사 3층 회의실 A", "agenda": "주간 실적 리뷰 / 경과채권 현황 공유", "attendees": "전원"},
    {"id": "m3", "datetime": d(8) + "T14:00",  "place": "화상회의(온라인)",   "agenda": "하반기 지역별 판매 목표 조정", "attendees": "최율아, 김도현, 박서연, 이준호"},
    {"id": "m4", "datetime": d(15) + "T10:00", "place": "본사 5층 대회의실", "agenda": "신모델 글로벌 론칭 준비 상황 점검", "attendees": "전원"},
]

# ---------------------------------------------------------------------------
# 경과채권 링크 6건
# ---------------------------------------------------------------------------
receivable_links = [
    {"id": "r1", "target": "유럽법인",   "fileName": "경과채권_유럽_" + MONTHS[-1] + ".xlsx",  "path": "\\\\efam\\해외영업\\경과채권\\유럽법인",   "updatedAt": d(-3)},
    {"id": "r2", "target": "미주법인",   "fileName": "경과채권_미주_" + MONTHS[-1] + ".xlsx",  "path": "\\\\efam\\해외영업\\경과채권\\미주법인",   "updatedAt": d(-3)},
    {"id": "r3", "target": "아시아법인", "fileName": "경과채권_아시아_" + MONTHS[-1] + ".xlsx", "path": "\\\\efam\\해외영업\\경과채권\\아시아법인", "updatedAt": d(-4)},
    {"id": "r4", "target": "직수출",     "fileName": "경과채권_직수출_" + MONTHS[-1] + ".xlsx", "path": "\\\\efam\\해외영업\\경과채권\\직수출",     "updatedAt": d(-4)},
    {"id": "r5", "target": "전사 요약",  "fileName": "경과채권_전사요약(구글시트)",             "path": "https://docs.google.com/spreadsheets/d/EXAMPLE-OVERDUE-SUMMARY/edit", "updatedAt": d(-1)},
    {"id": "r6", "target": "DL12 걸프스타 제너럴", "fileName": "걸프스타_채권회수계획.pdf",     "path": "\\\\efam\\해외영업\\경과채권\\중점관리\\DL12", "updatedAt": d(-6)},
]

# ---------------------------------------------------------------------------
# seed.js 출력
# ---------------------------------------------------------------------------
seed = {
    "users": USERS,
    "dealers": DEALERS,
    "dealerMetrics": metrics,
    "weeklyReports": weekly,
    "exchangeRates": rates,
    "trips": trips,
    "meetings": meetings,
    "receivableLinks": receivable_links,
    "notifyRecipients": [u["email"] for u in USERS] + ["gm.overseas@example.com"],
    "notifyLog": [],
}

seed_js = (
    "// 자동 생성 파일 — make_samples.py 실행으로 갱신됩니다. 직접 수정하지 마세요.\n"
    "// 생성일: " + datetime.now().strftime("%Y-%m-%d %H:%M") + "\n"
    "window.SEED_DATA = " + json.dumps(seed, ensure_ascii=False, indent=2) + ";\n"
)
os.makedirs(os.path.join(BASE, "js"), exist_ok=True)
with open(os.path.join(BASE, "js", "seed.js"), "w", encoding="utf-8") as f:
    f.write(seed_js)
print("js/seed.js 생성 완료 (딜러 %d, 실적 %d행, 주간업무 %d건)" % (len(DEALERS), len(metrics), len(weekly)))

# ---------------------------------------------------------------------------
# 엑셀 양식 / 샘플 생성
# ---------------------------------------------------------------------------
HEADERS = ["딜러코드", "딜러명", "국가", "기준월", "매출", "재고", "채권잔액", "경과채권", "계약조건"]
COL_WIDTHS = [10, 22, 12, 10, 12, 12, 12, 12, 34]
HEAD_FILL = PatternFill("solid", fgColor="1F3A5F")
HEAD_FONT = Font(color="FFFFFF", bold=True)
NOTE_FONT = Font(color="888888", size=9)


def build_workbook(with_sample):
    wb = Workbook()
    ws = wb.active
    ws.title = "딜러실적"
    for c, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    row = 2
    if with_sample:
        dl_map = {dl["code"]: dl for dl in DEALERS}
        for m in metrics:
            dl = dl_map[m["dealerCode"]]
            ws.cell(row=row, column=1, value=m["dealerCode"])
            ws.cell(row=row, column=2, value=dl["name"])
            ws.cell(row=row, column=3, value=dl["country"])
            ws.cell(row=row, column=4, value=m["month"])
            ws.cell(row=row, column=5, value=m["sales"])
            ws.cell(row=row, column=6, value=m["inventory"])
            ws.cell(row=row, column=7, value=m["receivable"])
            ws.cell(row=row, column=8, value=m["overdue"])
            ws.cell(row=row, column=9, value=m["terms"])
            row += 1
    guide = wb.create_sheet("작성안내")
    lines = [
        "딜러실적 업로드 양식 작성 안내",
        "",
        "1. '딜러실적' 시트에 한 행 = 딜러 1개 × 기준월 1개 로 입력합니다.",
        "2. 딜러코드: DL01~DL14 형식. 대시보드 딜러 구분 기준이므로 정확히 입력하세요.",
        "3. 기준월: YYYY-MM 형식 텍스트 (예: 2026-08).",
        "4. 매출/재고/채권잔액/경과채권: 숫자만 입력 (단위: USD 천불).",
        "5. 계약조건: 자유 텍스트 (예: T/T 30일, FOB).",
        "6. 같은 딜러코드+기준월 행을 다시 올리면 기존 값이 덮어쓰기(갱신)됩니다.",
        "7. 12개월 히스토리 추이를 보려면 딜러별로 최근 12개월 행을 채워 주세요.",
    ]
    for i, line in enumerate(lines, start=1):
        cell = guide.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=12)
    guide.column_dimensions["A"].width = 80
    note = ws.cell(row=1, column=len(HEADERS) + 2, value="※ 작성 방법은 '작성안내' 시트 참고")
    note.font = NOTE_FONT
    return wb


os.makedirs(os.path.join(BASE, "templates"), exist_ok=True)
build_workbook(False).save(os.path.join(BASE, "templates", "딜러실적_양식.xlsx"))
build_workbook(True).save(os.path.join(BASE, "templates", "딜러실적_샘플.xlsx"))
print("templates/딜러실적_양식.xlsx, templates/딜러실적_샘플.xlsx 생성 완료")
